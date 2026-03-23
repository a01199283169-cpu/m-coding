"""
엑셀 일괄 등록 화면
"""
import streamlit as st
from database import get_connection
from utils.code_gen import generate_item_code
from utils.excel_export import export_to_excel, get_excel_filename
from datetime import datetime
import io
from openpyxl import load_workbook


def get_dept_code_by_name(dept_name: str) -> str:
    """학과명으로 학과코드 조회"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT dept_code, college_code
        FROM departments
        WHERE dept_name = ? OR dept_name LIKE ?
    """, (dept_name, f"%{dept_name}%"))
    result = cursor.fetchone()
    conn.close()

    if result:
        return result['dept_code'], result['college_code']
    return None, None


def parse_excel_file(file_content: bytes) -> tuple:
    """
    엑셀 파일 파싱

    Returns:
        (success: bool, data: list or error_message: str)
    """
    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        data_list = []
        errors = []

        # 헤더 행 건너뛰기 (2행부터)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not any(row):
                continue

            try:
                # 엑셀 데이터 읽기 (단과대학 다음에 학과)
                # 번호(0), 단과대학(1), 학과(2), 자산구분(3), 카테고리(4), 품목코드(5), 품목명(6), 규격(7),
                # 수량(8), 단가(9), 총액(10), 입고일(11),
                # 납품업체(12), 보관위치(13), 내용연수(14), 폐기예정일(15),
                # 입력자(16), 입력일(17), 비고(18)

                item_name = row[6]
                category = row[4] if row[4] else '기타'
                spec = row[7]
                dept_name = row[2]
                quantity = int(row[8]) if row[8] else 1
                unit_price = int(row[9]) if row[9] else 0
                arrival_date = row[11]
                vendor = row[12]
                location = row[13]
                useful_life = int(row[14]) if row[14] else None
                registrant_name = row[16]
                note = row[18]

                # 필수 항목 검증
                if not item_name:
                    errors.append(f"행 {row_idx}: 품목명이 없습니다.")
                    continue

                if not dept_name:
                    errors.append(f"행 {row_idx}: 학과명이 없습니다.")
                    continue

                if not registrant_name:
                    errors.append(f"행 {row_idx}: 입력자 이름이 없습니다.")
                    continue

                # 학과명으로 학과코드 조회
                dept_code, college_code = get_dept_code_by_name(dept_name)
                if not dept_code:
                    errors.append(f"행 {row_idx}: 학과명 '{dept_name}'을 찾을 수 없습니다.")
                    continue

                # 입고일 처리
                if arrival_date:
                    if isinstance(arrival_date, datetime):
                        arrival_date_str = arrival_date.strftime('%Y-%m-%d')
                    else:
                        arrival_date_str = str(arrival_date)
                else:
                    arrival_date_str = datetime.now().strftime('%Y-%m-%d')

                # 구매일 = 입고일로 설정 (간소화)
                purchase_date_str = arrival_date_str

                # 총액 계산
                total_price = quantity * unit_price

                # 폐기예정일 계산
                disposal_date_str = None
                if useful_life and useful_life > 0:
                    from datetime import timedelta
                    arrival_dt = datetime.strptime(arrival_date_str, '%Y-%m-%d')
                    disposal_dt = arrival_dt + timedelta(days=useful_life * 365)
                    disposal_date_str = disposal_dt.strftime('%Y-%m-%d')

                item_data = {
                    'item_name': item_name,
                    'category': category,
                    'spec': spec,
                    'dept_code': dept_code,
                    'college_code': college_code,
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'total_price': total_price,
                    'purchase_date': purchase_date_str,
                    'arrival_date': arrival_date_str,
                    'vendor': vendor,
                    'asset_type': '교육용기자재',
                    'location': location,
                    'useful_life': useful_life,
                    'disposal_date': disposal_date_str,
                    'note': note,
                    'registrant_name': registrant_name
                }

                data_list.append(item_data)

            except Exception as e:
                errors.append(f"행 {row_idx}: {str(e)}")

        if errors:
            return False, "\n".join(errors)

        return True, data_list

    except Exception as e:
        return False, f"엑셀 파일 읽기 실패: {str(e)}"


def bulk_insert_equipment(data_list: list, username: str) -> tuple:
    """
    일괄 등록

    Returns:
        (success_count: int, error_count: int, errors: list)
    """
    conn = get_connection()
    cursor = conn.cursor()

    success_count = 0
    error_count = 0
    errors = []

    for idx, data in enumerate(data_list, start=1):
        try:
            # 품목코드 자동생성
            item_code = generate_item_code(data['dept_code'])

            # DB 삽입
            cursor.execute("""
                INSERT INTO equipment (
                    item_code, item_name, category, spec, dept_code, college_code,
                    quantity, unit_price, total_price,
                    purchase_date, arrival_date, vendor,
                    asset_type, location, useful_life, disposal_date,
                    note, registrant_name, registrant_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                item_code,
                data['item_name'],
                data.get('category'),
                data.get('spec'),
                data['dept_code'],
                data['college_code'],
                data['quantity'],
                data['unit_price'],
                data['total_price'],
                data['purchase_date'],
                data['arrival_date'],
                data.get('vendor'),
                data['asset_type'],
                data.get('location'),
                data.get('useful_life'),
                data.get('disposal_date'),
                data.get('note'),
                data['registrant_name'],
                username
            ))

            # 즉시 커밋하여 다음 품목코드 생성 시 중복 방지
            conn.commit()
            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"항목 {idx} ({data['item_name']}): {str(e)}")
            # 에러 발생 시 롤백
            conn.rollback()

    conn.close()

    return success_count, error_count, errors


def download_template():
    """엑셀 템플릿 다운로드"""
    # 샘플 데이터
    template_data = [
        {
            'item_code': '(자동생성)',
            'item_name': '노트북 컴퓨터',
            'category': 'IT기기',
            'spec': 'Intel i7, 16GB RAM',
            'college': '보건대학',
            'dept_name': '치위생학과',
            'asset_type': '교육용기자재',
            'quantity': 5,
            'unit_price': 1500000,
            'total_price': 7500000,
            'arrival_date': '2026-03-15',
            'vendor': 'ABC컴퓨터',
            'location': 'B동 301호',
            'useful_life': 5,
            'disposal_date': '2031-03-15',
            'registrant_name': '홍길동',
            'created_at': '2026-03-23',
            'note': '교육용'
        }
    ]

    excel_data = export_to_excel(template_data, "템플릿")
    return excel_data


def show_bulk_upload(user: dict):
    """엑셀 일괄 등록 메인 화면"""
    st.title("📤 엑셀 일괄 등록")

    st.info("""
    💡 **안내**
    - 엑셀 파일을 이용하여 여러 기자재를 한 번에 등록할 수 있습니다.
    - 템플릿을 다운로드하여 양식에 맞게 작성하세요.
    - 품목코드는 자동으로 생성되므로 입력하지 않아도 됩니다.
    - **사진 업로드는 등록 후 개별 수정에서 추가하세요.**
    """)

    st.markdown("---")

    # 탭 구성
    tab1, tab2 = st.tabs(["📥 템플릿 다운로드", "📤 파일 업로드"])

    with tab1:
        st.subheader("📥 엑셀 템플릿 다운로드")

        st.write("**템플릿 양식 (단과대학 다음에 학과)**")
        st.markdown("""
        | 번호 | 단과대학 | 학과 | 자산구분 | 카테고리 | 품목코드 | 품목명 | 규격 | 수량 | 단가 | 총액 | 입고일 | 납품업체 | 보관위치 | 내용연수 | 폐기예정일 | 입력자 | 입력일 | 비고 |
        |------|----------|------|----------|----------|----------|--------|------|------|------|------|--------|----------|----------|----------|------------|--------|--------|------|
        | 1 | 보건대학 | 치위생학과 | 교육용기자재 | IT기기 | (자동생성) | 노트북 컴퓨터 | Intel i7 | 5 | 1,500,000 | 7,500,000 | 2026-03-15 | ABC컴퓨터 | B동 301호 | 5 | 2031-03-15 | 홍길동 | 2026-03-23 | 교육용 |
        """)

        st.markdown("---")

        # 템플릿 다운로드 버튼
        template_data = download_template()
        st.download_button(
            label="⬇️ 템플릿 다운로드",
            data=template_data,
            file_name=f"기자재등록_템플릿_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        st.caption("※ 템플릿을 다운로드하여 데이터를 입력한 후 업로드하세요.")

    with tab2:
        st.subheader("📤 엑셀 파일 업로드")

        # 파일 업로드
        uploaded_file = st.file_uploader(
            "엑셀 파일 선택",
            type=['xlsx'],
            key="bulk_upload_file"
        )

        if uploaded_file:
            st.success(f"파일 선택됨: {uploaded_file.name}")

            # 미리보기 버튼
            col1, col2 = st.columns(2)

            with col1:
                if st.button("👀 미리보기", use_container_width=True):
                    # 파일 파싱
                    file_content = uploaded_file.getvalue()
                    success, result = parse_excel_file(file_content)

                    if success:
                        st.success(f"✅ {len(result)}건의 데이터를 확인했습니다.")

                        # 미리보기 테이블
                        preview_data = []
                        for idx, item in enumerate(result[:10], start=1):  # 최대 10개만 표시
                            preview_data.append({
                                '번호': idx,
                                '품목명': item['item_name'],
                                '카테고리': item['category'],
                                '학과': item['dept_code'],
                                '수량': item['quantity'],
                                '단가': f"{item['unit_price']:,}원",
                                '입력자': item['registrant_name']
                            })

                        st.table(preview_data)

                        if len(result) > 10:
                            st.info(f"※ 총 {len(result)}건 중 10건만 표시됩니다.")

                        # 세션에 저장
                        st.session_state.bulk_upload_data = result

                    else:
                        st.error("❌ 파일 파싱 실패")
                        st.error(result)

            with col2:
                if st.button("💾 일괄 등록", use_container_width=True, type="primary"):
                    if 'bulk_upload_data' not in st.session_state:
                        st.warning("먼저 미리보기를 실행하세요.")
                    else:
                        data_list = st.session_state.bulk_upload_data

                        with st.spinner(f"{len(data_list)}건 등록 중..."):
                            success_count, error_count, errors = bulk_insert_equipment(
                                data_list,
                                user['username']
                            )

                        # 결과 표시
                        if error_count == 0:
                            st.success(f"🎉 {success_count}건 모두 등록 완료!")
                            del st.session_state.bulk_upload_data
                        else:
                            st.warning(f"⚠️ {success_count}건 성공, {error_count}건 실패")

                            if errors:
                                with st.expander("❌ 오류 내역 보기"):
                                    for error in errors:
                                        st.error(error)

                        st.balloons()


if __name__ == "__main__":
    # 테스트용
    st.set_page_config(page_title="엑셀 일괄 등록", layout="wide")

    test_user = {
        'username': 'hlt_1',
        'role': 'college',
        'college_code': 'HLT',
        'college_name': '보건대학'
    }

    show_bulk_upload(test_user)
