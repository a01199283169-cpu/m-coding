"""
엑셀 다운로드 및 업로드 기능
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io


def export_to_excel(data: list, college_name: str = "전체") -> bytes:
    """
    기자재 데이터를 엑셀 파일로 변환

    Args:
        data: 기자재 데이터 리스트
        college_name: 단과대학명 (파일명용)

    Returns:
        엑셀 파일 바이트 데이터
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "기자재 현황"

    # 헤더 설정 (단과대학 다음에 학과)
    headers = [
        "번호", "단과대학", "학과", "자산구분", "카테고리", "품목코드", "자산코드", "품목명", "규격",
        "수량", "단가", "총액", "입고일",
        "납품업체", "보관위치", "예산부서", "내용연수", "폐기예정일",
        "입력자", "입력일", "비고"
    ]

    # 헤더 스타일
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 데이터 입력 (단과대학 다음에 학과)
    for row_idx, item in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # 번호
        ws.cell(row=row_idx, column=2, value=item.get('college', ''))  # 단과대학
        ws.cell(row=row_idx, column=3, value=item.get('dept_name', ''))  # 학과
        ws.cell(row=row_idx, column=4, value=item.get('asset_type', ''))  # 자산구분
        ws.cell(row=row_idx, column=5, value=item.get('category', ''))  # 카테고리
        ws.cell(row=row_idx, column=6, value=item.get('item_code', ''))  # 품목코드
        ws.cell(row=row_idx, column=7, value=item.get('asset_code', ''))  # 자산코드
        ws.cell(row=row_idx, column=8, value=item.get('item_name', ''))  # 품목명
        ws.cell(row=row_idx, column=9, value=item.get('spec', ''))  # 규격
        ws.cell(row=row_idx, column=10, value=item.get('quantity', 0))  # 수량

        # 단가 - 천단위 콤마
        cell_unit_price = ws.cell(row=row_idx, column=11, value=item.get('unit_price', 0))
        cell_unit_price.number_format = '#,##0'

        # 총액 - 천단위 콤마
        cell_total_price = ws.cell(row=row_idx, column=12, value=item.get('total_price', 0))
        cell_total_price.number_format = '#,##0'

        ws.cell(row=row_idx, column=13, value=item.get('arrival_date', ''))  # 입고일
        ws.cell(row=row_idx, column=14, value=item.get('vendor', ''))  # 납품업체
        ws.cell(row=row_idx, column=15, value=item.get('location', ''))  # 보관위치
        ws.cell(row=row_idx, column=16, value=item.get('budget_dept', ''))  # 예산부서
        ws.cell(row=row_idx, column=17, value=item.get('useful_life', ''))  # 내용연수
        ws.cell(row=row_idx, column=18, value=item.get('disposal_date', ''))  # 폐기예정일
        ws.cell(row=row_idx, column=19, value=item.get('registrant_name', ''))  # 입력자
        ws.cell(row=row_idx, column=20, value=item.get('created_at', ''))  # 입력일
        ws.cell(row=row_idx, column=21, value=item.get('note', ''))  # 비고

    # 열 너비 자동 조정
    column_widths = {
        'A': 6,  'B': 15, 'C': 25, 'D': 12, 'E': 20, 'F': 15, 'G': 15,
        'H': 20, 'I': 12, 'J': 8,  'K': 12, 'L': 12, 'M': 12,
        'N': 15, 'O': 15, 'P': 12, 'Q': 10, 'R': 12, 'S': 12, 'T': 12, 'U': 30
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 바이트로 변환
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file.getvalue()


def get_excel_filename(college_name: str = "전체") -> str:
    """엑셀 파일명 생성"""
    date_str = datetime.now().strftime('%Y%m%d')
    return f"기자재현황_{college_name}_{date_str}.xlsx"


def parse_excel_upload(file_content: bytes) -> tuple:
    """
    업로드된 엑셀 파일 파싱

    Args:
        file_content: 엑셀 파일 바이트 데이터

    Returns:
        (success: bool, data: list or error_message: str)
    """
    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        data_list = []
        errors = []

        # 헤더 행 건너뛰기
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 건너뛰기
            if not any(row):
                continue

            try:
                # 엑셀 데이터 → 딕셔너리 변환 (헤더 순서에 맞춤)
                # 번호(0), 단과대학(1), 학과(2), 자산구분(3), 카테고리(4), 품목코드(5), 자산코드(6),
                # 품목명(7), 규격(8), 수량(9), 단가(10), 총액(11), 입고일(12),
                # 납품업체(13), 보관위치(14), 예산부서(15), 내용연수(16), 폐기예정일(17),
                # 입력자(18), 입력일(19), 비고(20)
                item_data = {
                    'item_code': row[5],        # 품목코드
                    'asset_code': row[6],       # 자산코드
                    'item_name': row[7],        # 품목명
                    'spec': row[8],             # 규격
                    'dept_code': None,          # 학과코드 (학과명으로부터 조회 필요)
                    'dept_name': row[2],        # 학과명
                    'asset_type': row[3] if row[3] else '교육용기자재',
                    'category': row[4],         # 카테고리
                    'quantity': int(row[9]) if row[9] else 1,
                    'unit_price': int(row[10]) if row[10] else 0,
                    'total_price': int(row[11]) if row[11] else 0,
                    'arrival_date': row[12],    # 입고일
                    'vendor': row[13],          # 납품업체
                    'location': row[14],        # 보관위치
                    'budget_dept': row[15],     # 예산부서
                    'useful_life': int(row[16]) if row[16] else None,
                    'disposal_date': row[17],   # 폐기예정일
                    'registrant_name': row[18], # 입력자
                    'note': row[20]             # 비고
                }

                # 필수 항목 검증
                if not item_data['item_name']:
                    errors.append(f"행 {row_idx}: 품목명이 없습니다.")
                    continue

                if not item_data['dept_name']:
                    errors.append(f"행 {row_idx}: 학과명이 없습니다.")
                    continue

                data_list.append(item_data)

            except Exception as e:
                errors.append(f"행 {row_idx}: {str(e)}")

        if errors:
            return False, "\n".join(errors)

        return True, data_list

    except Exception as e:
        return False, f"엑셀 파일 읽기 실패: {str(e)}"


if __name__ == "__main__":
    # 테스트
    test_data = [
        {
            'item_code': 'DH-2026-0001',
            'item_name': '노트북 컴퓨터',
            'spec': 'Intel i7, 16GB RAM',
            'college': '보건대학',
            'dept_name': '치위생학과',
            'asset_type': '교육용기자재',
            'quantity': 5,
            'unit_price': 1500000,
            'total_price': 7500000,
            'arrival_date': '2026-03-15',
            'vendor': 'ABC컴퓨터',
            'location': 'B동 301호',
            'useful_life': 5,
            'disposal_date': '2031-03-15',
            'registrant_name': '홍길동',
            'created_at': '2026-03-23',
            'note': '교육용'
        }
    ]

    excel_data = export_to_excel(test_data, "보건대학")
    print(f"엑셀 파일 생성 완료: {len(excel_data)} bytes")
